#!/usr/bin/env python
"""
Custom Bracket Template Creator

This script creates a modified version of the invoice template with proper bracket placeholders
"""

import os
from pathlib import Path
import shutil
from openpyxl import load_workbook
import re
from openpyxl.utils import get_column_letter
from openpyxl.cell.cell import MergedCell

def create_custom_bracket_template(input_path, output_path):
    """
    Create a custom invoice template with proper bracket placeholders.

    Args:
        input_path: Path to the input XLSX file
        output_path: Path to save the converted file
    """
    print(f"Creating custom bracket template from {input_path}...")

    # First make a copy of the template
    shutil.copy2(input_path, output_path)

    # Load the copied workbook
    wb = load_workbook(output_path)

    # Define proper placeholder mappings - exact matches only to avoid over-replacement
    placeholders = {
        # Company information
        "Company Name": "[Company Name]",
        "Your Company": "[Company Name]",
        "Your Company Slogan": "[Company Slogan]",
        "Street Address": "[Company Address]",
        "City, ST ZIP Code": "[Company City]",
        "Phone": "[Company Phone]",
        "555-1234": "[Company Phone]",
        "Email": "[Company Email]",
        "Website": "[Company Website]",

        # Invoice information
        "INVOICE #": "INVOICE #: [Invoice Number]",
        "DATE": "DATE: [Invoice Date]",
        "DUE DATE": "DUE DATE: [Due Date]",
        "TERMS": "TERMS: [Payment Terms]",

        # Client information
        "Recipient Name": "[Client Name]",
        "Client Name": "[Client Name]",
        "Client Company": "[Client Company]",
        "Client Address": "[Client Address]",
        "Client Phone": "[Client Phone]",
        "Client Email": "[Client Email]",

        # Financial information
        "SUBTOTAL": "SUBTOTAL: [Subtotal]",
        "DISCOUNT": "DISCOUNT: [Discount Amount]",
        "TAX": "TAX: [Tax Amount]",
        "TOTAL": "TOTAL: [Total Amount]",
        "BALANCE DUE": "BALANCE DUE: [Total Amount]"
    }

    # Process each worksheet
    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        print(f"Processing sheet: {sheet_name}")

        # Process each cell
        for row in ws.iter_rows():
            for cell in row:
                if not isinstance(cell, MergedCell) and cell.value and isinstance(cell.value, str):
                    original_text = cell.value

                    # Check if the value already has bracket patterns
                    # If it does, we need to be careful not to nest brackets
                    if '[' in original_text and ']' in original_text:
                        # Remove any existing brackets from original text for comparison
                        clean_text = re.sub(r'\[.*?\]', '', original_text)

                        # If there's still text after removing brackets
                        if clean_text.strip():
                            new_text = original_text  # Start with original
                            for old_text, new_text_template in placeholders.items():
                                # Only replace exact words, not partial matches
                                pattern = r'\b' + re.escape(old_text) + r'\b'
                                if re.search(pattern, clean_text, re.IGNORECASE):
                                    new_text = re.sub(pattern, new_text_template, new_text, flags=re.IGNORECASE)
                        else:
                            # If it's just brackets, leave it as is
                            new_text = original_text
                    else:
                        # No brackets in the original text, safe to do replacements
                        new_text = original_text
                        for old_text, new_text_template in placeholders.items():
                            # Only replace exact words, not partial matches
                            pattern = r'\b' + re.escape(old_text) + r'\b'
                            if re.search(pattern, new_text, re.IGNORECASE):
                                new_text = re.sub(pattern, new_text_template, new_text, flags=re.IGNORECASE)

                    # Update cell if changes were made
                    if new_text != original_text:
                        print(f"Replacing in {sheet_name}: '{original_text}' -> '{new_text}'")
                        cell.value = new_text

        # Now directly set important placeholders in specific cells
        # This helps ensure all necessary placeholders are present
        set_direct_placeholders(ws)

    # Save the modified workbook
    wb.save(output_path)
    print(f"Saved custom bracket template to {output_path}")

def set_direct_placeholders(worksheet):
    """
    Set direct placeholder values in specific cells based on an analysis of the template structure.

    Args:
        worksheet: The worksheet to modify
    """
    # Company section (left upper corner)
    company_section = {
        "B2": "[Company Name]",
        "B3": "[Company Slogan]",
        "B4": "[Company Address]",
        "B5": "[Company City], [Company State] [Company Zip]",
        "B6": "Phone: [Company Phone]",
        "B7": "Email: [Company Email]",
        "B8": "Web: [Company Website]"
    }

    # Invoice info section (right upper corner)
    invoice_section = {
        "G3": "INVOICE #: [Invoice Number]",
        "G4": "DATE: [Invoice Date]",
        "G5": "DUE DATE: [Due Date]",
        "G6": "TERMS: [Payment Terms]"
    }

    # Client section (middle section)
    client_section = {
        "B11": "BILL TO:",
        "B12": "[Client Name]",
        "B13": "[Client Company]",
        "B14": "[Client Address]",
        "B15": "[Client City], [Client State] [Client Zip]",
        "B16": "Phone: [Client Phone]",
        "B17": "Email: [Client Email]"
    }

    # Invoice summary section (bottom right)
    summary_section = {
        "F31": "SUBTOTAL:",
        "G31": "[Subtotal]",
        "F32": "TAX:",
        "G32": "[Tax Amount]",
        "F33": "DISCOUNT:",
        "G33": "[Discount Amount]",
        "F34": "TOTAL:",
        "G34": "[Total Amount]"
    }

    # Combine all sections
    all_placeholders = {**company_section, **invoice_section, **client_section, **summary_section}

    # Set each placeholder
    for cell_ref, value in all_placeholders.items():
        cell = worksheet[cell_ref]

        # Skip merged cells
        if isinstance(cell, MergedCell):
            print(f"Skipping merged cell at {cell_ref}")
            continue

        print(f"Setting direct placeholder in {cell_ref}: '{value}'")
        cell.value = value

if __name__ == "__main__":
    template_dir = Path(__file__).resolve().parent
    input_template = template_dir / "InvoiceAssistantTemplate.xlsx"
    output_template = template_dir / "BracketInvoiceAssistantTemplate.xlsx"

    create_custom_bracket_template(input_template, output_template)